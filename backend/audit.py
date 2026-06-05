import csv
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, AuditLog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

audit_bp = Blueprint('audit', __name__, url_prefix='/api/audit')


@audit_bp.route('/', methods=['GET'])
@login_required
def get_logs():
    search = request.args.get('search', '')
    action = request.args.get('action')
    username = request.args.get('username')
    event_type = request.args.get('event_type')
    status = request.args.get('status')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    sort_by = request.args.get('sort_by', 'timestamp')
    sort_dir = request.args.get('sort_dir', 'desc')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))

    query = AuditLog.query

    if search:
        query = query.filter(
            db.or_(
                AuditLog.file_name.ilike(f'%{search}%'),
                AuditLog.username.ilike(f'%{search}%'),
                AuditLog.action.ilike(f'%{search}%'),
                AuditLog.details.ilike(f'%{search}%'),
            )
        )
    if action:
        query = query.filter(AuditLog.action.ilike(f'%{action}%'))
    if username:
        query = query.filter(AuditLog.username.ilike(f'%{username}%'))
    if event_type:
        query = query.filter_by(event_type=event_type)
    if status:
        query = query.filter_by(status=status)
    if date_from:
        try:
            query = query.filter(AuditLog.timestamp >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(AuditLog.timestamp <= datetime.fromisoformat(date_to))
        except ValueError:
            pass

    # Sorting
    col = getattr(AuditLog, sort_by, AuditLog.timestamp)
    if sort_dir == 'asc':
        query = query.order_by(col.asc())
    else:
        query = query.order_by(col.desc())

    total = query.count()
    logs = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'logs': [l.to_dict() for l in logs],
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
    }), 200


@audit_bp.route('/export', methods=['GET'])
@login_required
def export_logs():
    fmt = request.args.get('format', 'csv').lower()

    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).all()
    data = [l.to_dict() for l in logs]

    if fmt == 'csv':
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'audit_logs_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}.csv',
        )
    elif fmt in ('excel', 'xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Audit Logs'
        header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
        header_font = Font(color='00D4FF', bold=True)
        if data:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for row_idx, row in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(row.get(key, '') or ''))
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'audit_logs_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}.xlsx',
        )
    else:
        return jsonify({'error': 'Unsupported format. Use csv or excel'}), 400
