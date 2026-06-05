import csv
import io
import os
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, MonitoredFile, Alert, AuditLog, Report
from utils import log_audit, get_client_ip
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ReportLab imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


def _get_date_range(report_type, date_from=None, date_to=None):
    now = datetime.utcnow()
    if report_type == 'daily':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif report_type == 'weekly':
        start = now - timedelta(days=7)
        end = now
    elif report_type == 'monthly':
        start = now - timedelta(days=30)
        end = now
    elif report_type == 'custom' and date_from and date_to:
        start = datetime.fromisoformat(date_from)
        end = datetime.fromisoformat(date_to)
    else:
        start = now - timedelta(days=7)
        end = now
    return start, end


def _collect_stats(start, end):
    files_q = MonitoredFile.query.filter(MonitoredFile.upload_time.between(start, end))
    alerts_q = Alert.query.filter(Alert.timestamp.between(start, end))
    audit_q = AuditLog.query.filter(AuditLog.timestamp.between(start, end))

    stats = {
        'period_start': start.strftime('%Y-%m-%d %H:%M'),
        'period_end': end.strftime('%Y-%m-%d %H:%M'),
        'total_files': MonitoredFile.query.count(),
        'files_added': files_q.count(),
        'files_modified': MonitoredFile.query.filter_by(status='compromised').count(),
        'files_deleted': MonitoredFile.query.filter_by(status='deleted').count(),
        'total_alerts': alerts_q.count(),
        'active_alerts': alerts_q.filter_by(resolved=False).count(),
        'critical_alerts': alerts_q.filter_by(severity='CRITICAL').count(),
        'high_alerts': alerts_q.filter_by(severity='HIGH').count(),
        'medium_alerts': alerts_q.filter_by(severity='MEDIUM').count(),
        'low_alerts': alerts_q.filter_by(severity='LOW').count(),
        'audit_entries': audit_q.count(),
    }

    alert_rows = [a.to_dict() for a in alerts_q.order_by(Alert.timestamp.desc()).limit(50).all()]
    audit_rows = [l.to_dict() for l in audit_q.order_by(AuditLog.timestamp.desc()).limit(50).all()]
    file_rows = [f.to_dict() for f in MonitoredFile.query.order_by(MonitoredFile.upload_time.desc()).limit(50).all()]

    return stats, alert_rows, audit_rows, file_rows


def _generate_pdf(report_name, stats, alert_rows, audit_rows, file_rows, filepath):
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    # ── Colour palette ──
    DARK = colors.HexColor('#0a0e1a')
    BLUE = colors.HexColor('#0d6efd')
    CYAN = colors.HexColor('#00d4ff')
    GREEN = colors.HexColor('#00c853')
    RED = colors.HexColor('#ff1744')
    ORANGE = colors.HexColor('#ff6d00')
    GREY = colors.HexColor('#1e2a3a')
    LIGHT = colors.HexColor('#e0e6f0')

    header_style = ParagraphStyle('Header', parent=styles['Heading1'],
                                  fontSize=22, textColor=CYAN,
                                  alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'],
                               fontSize=9, textColor=LIGHT,
                               alignment=TA_CENTER, spaceAfter=8)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
                                   fontSize=12, textColor=CYAN,
                                   spaceBefore=14, spaceAfter=6)
    body_style = ParagraphStyle('Body', parent=styles['Normal'],
                                fontSize=9, textColor=LIGHT)

    # Header
    story.append(Paragraph('File Integrity Monitoring System Using Cybersecurity', header_style))
    story.append(Paragraph(f'Security Report — {report_name}', sub_style))
    story.append(Paragraph(
        f'Period: {stats["period_start"]} → {stats["period_end"]}  |  '
        f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}',
        sub_style))
    story.append(HRFlowable(width='100%', thickness=1, color=CYAN, spaceAfter=10))

    # Executive Summary
    story.append(Paragraph('Executive Summary', section_style))
    summary_data = [
        ['Metric', 'Count'],
        ['Total Files Monitored', str(stats['total_files'])],
        ['Files Added (Period)', str(stats['files_added'])],
        ['Files Modified / Compromised', str(stats['files_modified'])],
        ['Files Deleted', str(stats['files_deleted'])],
        ['Total Alerts (Period)', str(stats['total_alerts'])],
        ['Active Alerts', str(stats['active_alerts'])],
        ['Critical Alerts', str(stats['critical_alerts'])],
        ['High Severity Alerts', str(stats['high_alerts'])],
        ['Audit Log Entries', str(stats['audit_entries'])],
    ]
    t = Table(summary_data, colWidths=[10 * cm, 5 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), GREY),
        ('TEXTCOLOR', (0, 1), (-1, -1), LIGHT),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [GREY, colors.HexColor('#253447')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#334455')),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    # Alert Summary
    if alert_rows:
        story.append(Paragraph('Alert Details (Latest 50)', section_style))
        alert_table_data = [['#', 'File', 'Event', 'Severity', 'Time', 'Status']]
        for i, a in enumerate(alert_rows[:30], 1):
            sev = a.get('severity', '')
            sev_color = {'CRITICAL': '#ff1744', 'HIGH': '#ff6d00',
                         'MEDIUM': '#ffab00', 'LOW': '#00c853'}.get(sev, '#aaaaaa')
            alert_table_data.append([
                str(i),
                (a.get('file_name') or '')[:28],
                a.get('event_type', ''),
                sev,
                (a.get('timestamp') or '')[:16],
                'Resolved' if a.get('resolved') else 'Active',
            ])
        at = Table(alert_table_data, colWidths=[1*cm, 5*cm, 2.5*cm, 2.5*cm, 4*cm, 2.5*cm])
        at.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('BACKGROUND', (0, 1), (-1, -1), GREY),
            ('TEXTCOLOR', (0, 1), (-1, -1), LIGHT),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [GREY, colors.HexColor('#253447')]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#334455')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(at)
        story.append(Spacer(1, 14))

    # Monitored Files
    if file_rows:
        story.append(Paragraph('Monitored Files (Latest 30)', section_style))
        file_table_data = [['#', 'File Name', 'Type', 'Size (B)', 'Status', 'Upload Time']]
        for i, f in enumerate(file_rows[:30], 1):
            file_table_data.append([
                str(i),
                (f.get('file_name') or '')[:30],
                f.get('file_type', ''),
                str(f.get('file_size', 0)),
                f.get('status', ''),
                (f.get('upload_time') or '')[:16],
            ])
        ft = Table(file_table_data, colWidths=[0.8*cm, 5.5*cm, 1.8*cm, 2*cm, 2.2*cm, 4*cm])
        ft.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('BACKGROUND', (0, 1), (-1, -1), GREY),
            ('TEXTCOLOR', (0, 1), (-1, -1), LIGHT),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [GREY, colors.HexColor('#253447')]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#334455')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(ft)

    # Footer
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width='100%', thickness=0.5, color=GREY))
    story.append(Paragraph(
        'Generated by File Integrity Monitoring System Using Cybersecurity | Cybersecurity Report | CONFIDENTIAL',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7,
                       textColor=colors.HexColor('#556677'), alignment=TA_CENTER)
    ))

    doc.build(story)


def _generate_excel(report_name, stats, alert_rows, audit_rows, file_rows, filepath):
    header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
    header_font = Font(color='00D4FF', bold=True)

    def _write_sheet(wb, title, rows):
        ws = wb.create_sheet(title=title)
        if not rows:
            return
        headers = list(rows[0].keys())
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(headers, 1):
                ws.cell(row=ri, column=ci, value=str(row.get(key, '') or ''))
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb = Workbook()
    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    summary_rows = [{'Metric': k.replace('_', ' ').title(), 'Value': str(v)} for k, v in stats.items()]
    for ci, h in enumerate(['Metric', 'Value'], 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(summary_rows, 2):
        ws_sum.cell(row=ri, column=1, value=row['Metric'])
        ws_sum.cell(row=ri, column=2, value=row['Value'])
    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 20

    if alert_rows:
        _write_sheet(wb, 'Alerts', alert_rows)
    if file_rows:
        _write_sheet(wb, 'Files', file_rows)
    if audit_rows:
        _write_sheet(wb, 'Audit Logs', audit_rows)

    wb.save(filepath)


def _generate_csv(stats, alert_rows, file_rows, filepath):
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['=== SECURITY REPORT SUMMARY ==='])
        writer.writerow(['Metric', 'Value'])
        for k, v in stats.items():
            writer.writerow([k, v])
        writer.writerow([])
        if alert_rows:
            writer.writerow(['=== ALERTS ==='])
            writer.writerow(list(alert_rows[0].keys()))
            for row in alert_rows:
                writer.writerow(list(row.values()))
        writer.writerow([])
        if file_rows:
            writer.writerow(['=== FILES ==='])
            writer.writerow(list(file_rows[0].keys()))
            for row in file_rows:
                writer.writerow(list(row.values()))


@reports_bp.route('/generate', methods=['POST'])
@login_required
def generate_report():
    data = request.get_json()
    report_type = data.get('report_type', 'weekly')
    fmt = data.get('format', 'pdf').lower()
    date_from = data.get('date_from')
    date_to = data.get('date_to')
    ip = get_client_ip(request)

    from flask import current_app
    reports_folder = current_app.config['REPORTS_FOLDER']
    os.makedirs(reports_folder, exist_ok=True)

    start, end = _get_date_range(report_type, date_from, date_to)
    stats, alert_rows, audit_rows, file_rows = _collect_stats(start, end)

    ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    report_name = f'{report_type.capitalize()} Security Report'
    ext_map = {'pdf': 'pdf', 'excel': 'xlsx', 'csv': 'csv'}
    ext = ext_map.get(fmt, 'pdf')
    filename = f'report_{report_type}_{ts}.{ext}'
    filepath = os.path.join(reports_folder, filename)

    try:
        if fmt == 'pdf':
            _generate_pdf(report_name, stats, alert_rows, audit_rows, file_rows, filepath)
        elif fmt == 'excel':
            _generate_excel(report_name, stats, alert_rows, audit_rows, file_rows, filepath)
        elif fmt == 'csv':
            _generate_csv(stats, alert_rows, file_rows, filepath)
        else:
            return jsonify({'error': 'Unsupported format'}), 400
    except Exception as e:
        return jsonify({'error': f'Report generation failed: {str(e)}'}), 500

    # Save to Reports table
    report = Report(
        report_name=report_name,
        report_type=report_type,
        format=fmt,
        file_path=filepath,
        generated_by=current_user.id,
        date_from=start,
        date_to=end,
    )
    db.session.add(report)
    db.session.commit()

    log_audit(user_id=current_user.id, username=current_user.username,
              action='REPORT_GENERATED', details=f'{report_name} ({fmt})',
              ip_address=ip)

    return jsonify({
        'message': 'Report generated successfully',
        'report': report.to_dict(),
        'filename': filename,
        'stats': stats,
    }), 201


@reports_bp.route('/download/<filename>', methods=['GET'])
@login_required
def download_report(filename):
    from flask import current_app
    reports_folder = current_app.config['REPORTS_FOLDER']
    safe_path = os.path.join(reports_folder, os.path.basename(filename))

    if not os.path.exists(safe_path):
        return jsonify({'error': 'Report file not found'}), 404

    mime_map = {
        '.pdf': 'application/pdf',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.csv': 'text/csv',
    }
    ext = os.path.splitext(filename)[1].lower()
    mimetype = mime_map.get(ext, 'application/octet-stream')

    return send_file(safe_path, mimetype=mimetype, as_attachment=True,
                     download_name=filename)


@reports_bp.route('/list', methods=['GET'])
@login_required
def list_reports():
    reports = Report.query.order_by(Report.generated_at.desc()).limit(50).all()
    return jsonify({'reports': [r.to_dict() for r in reports]}), 200
